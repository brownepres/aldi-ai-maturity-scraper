import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import numpy as np
import seaborn as sns

colors = [
    '#F9E0E0',  # Nagyon halvány pasztell rózsaszín
    '#E7F9E7',  # Nagyon halvány pasztell zöld
    '#FFFCE1',  # Nagyon halvány pasztell sárga
    '#E9D9FF',  # Nagyon halvány pasztell lila
    '#E7F6FF',  # Nagyon halvány pasztell kék
    '#FFF1F4',  # Nagyon halvány pasztell barack
    '#D9F2FF',  # Nagyon halvány pasztell égkék
    '#FBEAEA',  # Nagyon halvány pasztell piros
    '#F8E8D3',  # Nagyon halvány pasztell homokszín
    '#F6E8FF'   # Nagyon halvány pasztell levendula
]

def makeVisuals(data):
    df = data
    data.replace('', np.nan, inplace = True)
    data = data.dropna()
    data.reset_index(drop=True, inplace=True)
    writer = pd.ExcelWriter("output_visual.xlsx", engine='openpyxl')
    df.to_excel(writer, sheet_name='data', index=False)

    pd.DataFrame().to_excel(writer, sheet_name='visualisation', index=False)

    #pie chart for news
    news_counts = data['Number of news'].value_counts().sort_index()
    
    fig, ax = plt.subplots(figsize=(6, 6))  
    ax.pie(news_counts, labels=news_counts.index, autopct='%1.1f%%', startangle=90, colors=colors)
    ax.axis('equal')  # A kör alak megőrzése
    ax.set_title("Distribution of 'Number of News'", fontsize=16)
    plt.savefig("pie_chart.png", format="png")
    plt.close() 

    #put pie chart on the sheet
    img_pie = Image("pie_chart.png")
    sheet = writer.book["visualisation"]
    sheet.add_image(img_pie, 'B2')


    #create histogram for average social media sentiment
    fig, ax = plt.subplots(figsize=(12, 8)) 
    companies = [company.split()[0] for company in data['Company']] 
    sentiment_vals = data['Average social media sentiment'].fillna(0).to_list()
    colors_for_hist = ['#D1ECFF' if val < 0 else '#F0D8D8' for val in sentiment_vals]
    ax.bar(companies, data['Average social media sentiment'], color=colors_for_hist)
    ax.set_xlabel("Company")
    ax.set_ylabel("Average Social Media Sentiment")
    ax.set_title("Average Social Media Sentiment by Company", fontsize=16)
    ax.axhline(0, color='black', linewidth=1)
    plt.xticks(rotation=90)

    #save diagram
    plt.savefig("bar_chart.png", format="png")
    plt.close() 

    #put histogram on the sheet
    img_bar = Image("bar_chart.png")
    sheet.add_image(img_bar, 'B26') 


    #create histogram for average news sentiment
    fig, ax = plt.subplots(figsize=(12, 8))  # Magasabb diagram a tengelyfeliratoknak
    companies = [company.split()[0] for company in data['Company']]  # Az első szó kiválasztása
    colors_for_hist = ['#D1ECFF' if val < 0 else '#F0D8D8' for val in data['Average news sentiment']]  # Pasztell rózsaszín és menta zöld

    ax.bar(companies, data['Average news sentiment'], color=colors_for_hist)
    ax.set_xlabel("Company")
    ax.set_ylabel("Average News Sentiment")
    ax.set_title("Average News Sentiment by Company", fontsize=16)
    ax.axhline(0, color='black', linewidth=1)  # 0-ás vonal a tengelyen (negatív értékek is megjelenhetnek)
    plt.xticks(rotation=90)  # A tengelyfeliratok elforgatása függőlegesen

    plt.savefig("bar_chart_news_sentiment.png", format="png")
    plt.close() 

    img_bar_news_sentiment = Image("bar_chart_news_sentiment.png")
    sheet.add_image(img_bar_news_sentiment, 'P26')

    #create merged cells in excel
    def create_merged_cell(sheet, merge_range, cell_address, text, color):
        
        # Összevont cellák létrehozása
        sheet.merge_cells(merge_range)

        # Cella kiválasztása
        cell = sheet[cell_address]
        cell.value = text  # Szöveg beállítása

        # A szöveg középre igazítása és félkövér betűtípus beállítása
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)

        # Egyéni háttérszín beállítása
        custom_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.fill = custom_fill

    create_merged_cell(sheet, 'Q2:S3', 'Q2', 'Average', "D1ECFF")  # Világoskék háttér
    try:
        create_merged_cell(sheet, 'U2:W3', 'U2', 'ALDI', "F0D8D8")   
    except:
        print("No ALDI found ")
    create_merged_cell(sheet, 'U5:W5', 'U5', 'Number of News', "F0D8D8")  # Halvány rózsaszín
    create_merged_cell(sheet, 'Q5:S5', 'Q5', 'Number of News', "D1ECFF")  
    create_merged_cell(sheet, 'Q6:S7', 'Q6', str(round(np.mean(data['Number of news']), 2)), "F9F9F9")  
    try:
        create_merged_cell(sheet, 'U6:W7', 'U6', str(round(data.loc[data['Company'] == 'ALDI', 'Number of news'].values[0], 2)), "F9F9F9")
    except:
        print("No aldi found")
    create_merged_cell(sheet, 'U9:W9', 'U9', 'Average news sentiment', "F0D8D8")  # Halvány rózsaszín
    create_merged_cell(sheet, 'Q9:S9', 'Q9', 'Average news sentiment', "D1ECFF")  
    create_merged_cell(sheet, 'Q10:S11', 'Q10', str(round(np.mean(data['Average news sentiment']), 2)), "F9F9F9")  
    try:
        create_merged_cell(sheet, 'U10:W11', 'U10', str(round(data.loc[data['Company'] == 'ALDI', 'Average news sentiment'].values[0], 2)), "F9F9F9")
    except:
        print("No aldi found")
    create_merged_cell(sheet, 'U13:W13', 'U13', 'Number of AI key words', "F0D8D8")  # Halvány rózsaszín
    create_merged_cell(sheet, 'Q13:S13', 'Q13', 'Number of AI key words', "D1ECFF")  
    create_merged_cell(sheet, 'Q14:S15', 'Q14', str(round(np.mean(data['Number of AI key words']), 2)), "F9F9F9")  
    try:
        create_merged_cell(sheet, 'U14:W15', 'U14', str(round(data.loc[data['Company'] == 'ALDI', 'Number of AI key words'].values[0], 2)), "F9F9F9")
    except:
        print("No aldi found")
    create_merged_cell(sheet, 'U17:W17', 'U17', 'Average social media sentiment', "F0D8D8")  # Halvány rózsaszín
    create_merged_cell(sheet, 'Q17:S17', 'Q17', 'Average social media sentiment', "D1ECFF")  
    create_merged_cell(sheet, 'Q18:S19', 'Q18', str(round(np.mean(data['Average social media sentiment']), 2)), "F9F9F9")  
    try:
        create_merged_cell(sheet, 'U18:W19', 'U18', str(round(data.loc[data['Company'] == 'ALDI', 'Average social media sentiment'].values[0], 2)), "F9F9F9")
    except:
        print("No aldi found")
    create_merged_cell(sheet, 'U21:W21', 'U21', 'AI buzz score', "F0D8D8")  # Halvány rózsaszín
    create_merged_cell(sheet, 'Q21:S21', 'Q21', 'AI buzz score', "D1ECFF")  
    create_merged_cell(sheet, 'Q22:S23', 'Q22', str(round(np.mean(data['AI buzz score']), 2)), "F9F9F9")  
    try:
        create_merged_cell(sheet, 'U22:W23', 'U22', str(round(data.loc[data['Company'] == 'ALDI', 'AI buzz score'].values[0], 2)), "F9F9F9")
    except:
        print("No aldi found")

    #creating visuals for financial data
    #preparation

    value_vars = [col for col in data.columns if 'profit_rate' in col or 'profitrate' in col or 'profit' in col.lower()]
    id_vars = [col for col in data.columns if col not in value_vars]
    melted = data.melt(id_vars=id_vars, value_vars=value_vars, var_name='metric_year', value_name='profit rate')
    melted['Year'] = melted['metric_year'].str.extract(r'(\d{4})')
    melted['metric'] = melted['metric_year'].str.extract(r'\d{4}[_]?(.+)').iloc[:, 0].str.replace('_', ' ').str.lower()
    melted = melted[melted['metric'].str.contains('profit rate', case=False)]
    melted['Company'] = melted['Company'].astype(str)

    colors_stronger = [
        '#B3DDF2',  # Világos pasztellkék
        '#7FC9F1',  # Középkék, világos, de élénk
        '#4FB0E6',  # Erőteljesebb kék
        '#2C98D1'   # Mélyebb, de még mindig világosabb kék
    ]
    plt.figure(figsize=(14, 8))
    sns.barplot(
        x="Company",
        y="profit rate",
        hue="Year",
        data=melted,
        palette=colors_stronger
    )

    plt.title("Profit Rate Annual Change by Company", fontsize=16)
    plt.xlabel("Company", fontsize=14)
    plt.ylabel("Profit Rate (%)", fontsize=14)
    plt.xticks(rotation=90)
    plt.legend(title='Year', loc='upper left')
    plt.tight_layout()
    plt.savefig("profit_rate_by_company_year.png", format="png")
    plt.close()

    # Insert image into Excel (assuming 'writer' and 'create_merged_cell' are defined)
    sheet = writer.book["visualisation"]
    img_profit_rate = Image("profit_rate_by_company_year.png")
    sheet.add_image(img_profit_rate, 'B80')

    # Plot: AI Buzz Score vs Profit Rate
    # Assumes you still have a column 'AI buzz score' and some 'profit rate' column
    # For simplicity, use the latest year's profit rate
    latest_year = melted['Year'].max()
    merged = data.copy()
    merged['profit rate'] = melted[melted['Year'] == latest_year].set_index('Company')['profit rate'].reindex(data['Company'].values).values

    plt.figure(figsize=(10, 6))
    x = merged['AI buzz score']
    y = merged['profit rate']
    plt.scatter(x, y, alpha=0.5, c=x, cmap='viridis', edgecolors='w', linewidth=0.5)

    plt.title("AI Buzz Score vs Profit Rate", fontsize=16)
    plt.xlabel("AI Buzz Score", fontsize=14)
    plt.ylabel("Profit Rate (%)", fontsize=14)
    plt.tight_layout()
    plt.savefig("ai_buzz_vs_profit_rate.png", format="png")
    plt.close()

    sheet = writer.book["visualisation"]
    img_ai_buzz_vs_profit_rate = Image("ai_buzz_vs_profit_rate.png")
    sheet.add_image(img_ai_buzz_vs_profit_rate, 'R80')

    # Correlation and merged cell update
    create_merged_cell(sheet, 'T84:V88', 'T84', 'Correlation between AI buzz score and Profit Rate', "D1ECFF")
    cell = sheet['T84']
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D1ECFF", end_color="D1ECFF", fill_type="solid")

    correlation = x.corr(y)
    create_merged_cell(sheet, 'W84:Y88', 'W84', str(round(correlation, 2)), "F9F9F9")

    writer.close()