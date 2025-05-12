# -*- coding: utf-8 -*-
"""
Created on Mon May 12 23:50:15 2025

@author: Gergő
"""

import feedparser
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
from collections import Counter
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pandas as pd
import requests
from bs4 import BeautifulSoup

import re
import requests
from bs4 import BeautifulSoup
from collections import Counter
import nltk

from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer, PorterStemmer
from nltk.tokenize import word_tokenize

from sklearn.feature_extraction.text import TfidfVectorizer

from collections import Counter
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

nltk.download("stopwords")
nltk.download("punkt")
nltk.download("wordnet")
nltk.download("omw-1.4")

stop_words = set(stopwords.words("english"))
lemmatizer = WordNetLemmatizer()
stemmer = PorterStemmer()


def preprocess_tokens(text):
    tokens = word_tokenize(text.lower())
    tokens = [word for word in tokens if word.isalpha() and word not in stop_words]
    tokens = [stemmer.stem(lemmatizer.lemmatize(word)) for word in tokens]
    return tokens


def celonis():
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)

    urls = [
        "https://www.celonis.com/press/partner-press-releases/"
    ]

    all_articles = []

    
    for url in urls:
        driver.get(url)

        # Cookie elfogadása, ha van
        try:
            WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler"))
            ).click()
        except:
            pass

        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CLASS_NAME, "ems-press-item"))
            )

            articles = driver.find_elements(By.CLASS_NAME, "ems-press-item")
            counter = 0
            for article in articles:
                if counter < 40:
                    try:
                        title_elem = article.find_element(By.CLASS_NAME, "ems-press-item__title")
                        title = title_elem.text.strip()
                        relative_link = article.get_attribute("href")
                        full_link = f"https://www.celonis.com{relative_link}" if relative_link.startswith("/") else relative_link
    
                        all_articles.append({
                            "title": title,
                            "link": full_link
                        })
    
                    except Exception as e:
                        print(f"Elem feldolgozási hiba: {e}")
                        continue
                    counter = counter + 1

        except Exception as e:
            print(f"Cikkek betöltési hiba az oldalon: {url} -> {e}")

    for article in all_articles:
        print(f"Cím: {article['title']}")
        print(f"Link: {article['link']}\n")

    driver.quit()
    return all_articles




def news(pages=20):
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)

    base_url = "https://www.marketresearchfuture.com/news"
    all_news = []

    for page_num in range(1, pages + 1):
        url = f"{base_url}?page={page_num}"
        print(f"Oldal betöltése: {url}")
        driver.get(url)

        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CLASS_NAME, "layout"))
            )

            articles = driver.find_elements(By.CLASS_NAME, "layout")

            for article in articles:
                try:
                    title_elem = article.find_element(By.TAG_NAME, "h6").find_element(By.TAG_NAME, "a")
                    title = title_elem.text.strip()
                    relative_link = title_elem.get_attribute("href")

                    # Teljes link
                    full_link = relative_link if relative_link.startswith("http") else base_url + relative_link

                    if title and full_link:
                        all_news.append({
                            "title": title,
                            "link": full_link
                        })

                except Exception as e:
                    print(f"Hiba a cikk feldolgozásakor: {e}")
                    continue

        except Exception as e:
            print(f"Hiba az oldal betöltésekor ({url}): {e}")

    driver.quit()

    # Kiíratás
    for news in all_news:
        print(f"Cím: {news['title']}")
        print(f"Link: {news['link']}\n")

    return all_news


def marketresearchfuture(base_url, pages=20):
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)

    root_url = "https://www.marketresearchfuture.com"
    all_reports = []

    for page_num in range(1, pages + 1):
        url = f"{base_url}?page={page_num}"
        print(f"Oldal betöltése: {url}")
        driver.get(url)

        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CLASS_NAME, "iiborder"))
            )

            report_blocks = driver.find_elements(By.CLASS_NAME, "iiborder")

            for block in report_blocks:
                try:
                    # Első link és href (ez az oldal linkje)
                    report_link_tag = block.find_element(By.TAG_NAME, "a")
                    relative_link = report_link_tag.get_attribute("href")
                    full_link = (
                        root_url + relative_link
                        if relative_link.startswith("/")
                        else relative_link
                    )

                    # Cím kinyerése a span vagy strong-ból
                    try:
                        title = block.find_element(By.CLASS_NAME, "report-icon").text.strip()
                    except:
                        title = block.find_element(By.TAG_NAME, "strong").text.strip()

                    # Dátum
                    try:
                        date = block.find_element(By.CLASS_NAME, "date").text.strip()
                    except:
                        date = ""

                    all_reports.append({
                        "title": title,
                        "link": full_link,
                        "date": date
                    })

                except Exception as e:
                    print(f"Hiba egy elem feldolgozásakor: {e}")
                    continue

        except Exception as e:
            print(f"Hiba az oldal betöltésekor ({url}): {e}")

    driver.quit()

    for report in all_reports:
        print(f"Cím: {report['title']}")
        print(f"Dátum: {report['date']}")
        print(f"Link: {report['link']}\n")

    return all_reports

def sstonework():
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)

    urls = [
        "https://www.ssonetwork.com/articles"
    ]

    all_articles = []
    base_url = "https://www.ssonetwork.com"

    for url in urls:
        driver.get(url)

        # Töltsük be az oldalon a további elemeket
        for i in range(5):
            try:
                load_more_btn = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.CLASS_NAME, "btn-load-more"))
                )
                driver.execute_script("arguments[0].click();", load_more_btn)
                time.sleep(2)  # Várjunk, míg betöltődnek az új elemek
            except Exception as e:
                print(f"Nem sikerült a {i+1}. Load more kattintás: {e}")
                break

        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CLASS_NAME, "post-item"))
            )

            article_divs = driver.find_elements(By.CLASS_NAME, "post-item")

            for div in article_divs:
                try:
                    title_elem = div.find_element(By.CLASS_NAME, "article-title").find_element(By.TAG_NAME, "a")
                    title = title_elem.text.strip()
                    relative_link = title_elem.get_attribute("href")
                    full_link = relative_link if relative_link.startswith("http") else base_url + relative_link

                    date_span = div.find_element(By.CLASS_NAME, "text-date").find_element(By.TAG_NAME, "span")
                    date = date_span.text.strip()

                    all_articles.append({
                        "title": title,
                        "date": date,
                        "link": full_link
                    })

                except Exception as e:
                    print(f"Hiba egy elem feldolgozásakor: {e}")
                    continue

        except Exception as e:
            print(f"Hiba a cikkek betöltésekor ezen az oldalon: {url} -> {e}")

    for article in all_articles:
        print(f"Cím: {article['title']}")
        print(f"Dátum: {article['date']}")
        print(f"Link: {article['link']}\n")

    driver.quit()
    return all_articles

def thinkhdi():
    options = webdriver.ChromeOptions()
    #options.add_argument("--headless")  # opcionális: ne nyissa meg a böngészőablakot
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)

    base_url = "https://www.thinkhdi.com/library/supportworld/metrics?pg="
    all_articles = []

    try:
        for page in range(1, 21):
            url = base_url + str(page)
            driver.get(url)

            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located((By.CLASS_NAME, "TaggedContent-Item"))
                )
            except Exception as e:
                print(f"Nem található tartalom az oldalon ({url}): {e}")
                continue

            articles = driver.find_elements(By.CLASS_NAME, "TaggedContent-Item")

            for article in articles:
                try:
                    title_elem = article.find_element(By.CLASS_NAME, "TaggedContent-Title").find_element(By.TAG_NAME, "a")
                    title = title_elem.text.strip()
                    link = title_elem.get_attribute("href")

                    all_articles.append({
                        "title": title,
                        "link": link
                    })

                except Exception as e:
                    print(f"Hiba egy elem feldolgozásakor: {e}")
                    continue

    finally:
        driver.quit()

    for art in all_articles:
        print(f"Cím: {art['title']}")
        print(f"Link: {art['link']}\n")

    return all_articles


def servicenow():
    # WebDriver beállítása
    driver = webdriver.Chrome()

    # URL, amelyet scrapelni szeretnél
    url = "https://www.servicenow.com/research/publication.html"

    # Oldal betöltése
    driver.get(url)

    # Várakozás, hogy a JavaScript betöltse az adatokat
    time.sleep(5)

    # Cikkek keresése
    publications = driver.find_elements(By.CLASS_NAME, 'pub-list-item')

    # Változó a cikkek mentésére
    publications_data = []

    # Cikkek címének és PDF linkjének kinyerése
    for pub in publications:
        title = pub.find_element(By.CSS_SELECTOR, 'a[style="font-weight: bold;"]')
        article_link_element = pub.find_element(By.CSS_SELECTOR, 'a[style*="font-weight: bold"]')
        pdf_link = article_link_element.get_attribute('href')
        
 
        print(pdf_link)
        
        # Ha mindkét adat létezik, hozzáadjuk a listához
        if title and pdf_link:
            publications_data.append({
                'title': title.text,
                'link': pdf_link
            })

    # A kinyert adatokat egy változóban tároljuk
    # Kinyomtathatjuk a változó tartalmát, ha szükséges
    for publication in publications_data:
        print(f"Cikk cím: {publication['title']}")
        print(f"PDF link: {publication['link']}\n")
        
    # WebDriver bezárása
    driver.quit()

    return publications_data


def clean_articles(articles):
    raw_data = []
    processed_texts = []
    token_lists = []

    for article in articles:
        time.sleep(10)
        title = article.get("title", "")
        link = article.get("link", "")
        content = ""

        if link:
            try:
                response = requests.get(link, timeout=10)
                if response.status_code == 200 or response.status_code == 403:
                    response.encoding = response.apparent_encoding
                    soup = BeautifulSoup(response.text, "html.parser")
                    print(soup)
                    paragraphs = soup.find_all("p")
                    print(paragraphs)
                    content = " ".join(p.get_text(strip=True) for p in paragraphs[:10])
            except Exception as e:
                print(f"Hiba a tartalom lekérésekor ({link}): {e}")
                content = "[Hiba a tartalom lekérésekor]"

        tokens = preprocess_tokens(content)
        token_lists.append(tokens)
        processed_texts.append(" ".join(tokens))

        raw_data.append({
            "Cikk címe": title,
            "Cikk linkje": link,
            "Cikk tartalma": content
        })

    # Top 10 common words (after stemming + lemmatization)
    for i, article in enumerate(raw_data):
        counter = Counter(token_lists[i])
        top10_common = [word for word, _ in counter.most_common(10)]
        article["Top 10 szó"] = ", ".join(top10_common)

    # TF-IDF számítás
    vectorizer = TfidfVectorizer()
    tfidf_matrix = vectorizer.fit_transform(processed_texts)
    feature_names = vectorizer.get_feature_names_out()

    for i, article in enumerate(raw_data):
        tfidf_vector = tfidf_matrix[i]
        tfidf_scores = zip(tfidf_vector.indices, tfidf_vector.data)
        sorted_words = sorted(tfidf_scores, key=lambda x: x[1], reverse=True)
        top_words = [feature_names[idx] for idx, _ in sorted_words[:10]]
        article["Top 10 TFIDF szó"] = ", ".join(top_words)

    return raw_data


def clean_articles_servicenow(articles):
    raw_data = []
    processed_texts = []
    token_lists = []
    counter = 0

    
    for article in articles:
        if counter < 10:
            title = article.get("title", "")
            link = article.get("link", "")
            content = ""
    
            if link:
                driver = webdriver.Chrome()
                driver.get(link)
                time.sleep(5)
    
                soup = BeautifulSoup(driver.page_source, "html.parser")
                paragraphs = soup.find_all("p")
                content = " ".join(p.get_text(strip=True) for p in paragraphs[3:4])
                print(content)
                driver.quit()
                
            tokens = preprocess_tokens(content)
            token_lists.append(tokens)
            processed_texts.append(" ".join(tokens))
    
            raw_data.append({
                "Cikk címe": title,
                "Cikk linkje": link,
                "Cikk tartalma": content
            })
            #counter = counter + 1

    # Top 10 common words (after stemming + lemmatization)
    for i, article in enumerate(raw_data):
        counter = Counter(token_lists[i])
        top10_common = [word for word, _ in counter.most_common(10)]
        article["Top 10 szó"] = ", ".join(top10_common)

    # TF-IDF számítás
    vectorizer = TfidfVectorizer()
    tfidf_matrix = vectorizer.fit_transform(processed_texts)
    feature_names = vectorizer.get_feature_names_out()

    for i, article in enumerate(raw_data):
        tfidf_vector = tfidf_matrix[i]
        tfidf_scores = zip(tfidf_vector.indices, tfidf_vector.data)
        sorted_words = sorted(tfidf_scores, key=lambda x: x[1], reverse=True)
        top_words = [feature_names[idx] for idx, _ in sorted_words[:10]]
        article["Top 10 TFIDF szó"] = ", ".join(top_words)

    return raw_data


from bs4 import BeautifulSoup
from collections import Counter
from sklearn.feature_extraction.text import TfidfVectorizer
from selenium import webdriver
import time
import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize

nltk.download('punkt')
nltk.download('stopwords')

def preprocess_tokens(text):
    tokens = word_tokenize(text.lower())
    stop_words = set(stopwords.words("english"))
    stop_words.update(["celonis"])  # "celonis" kizárása
    filtered = [w for w in tokens if w.isalpha() and w not in stop_words]
    return filtered

def clean_articles_celonis(articles):
    raw_data = []
    processed_texts = []
    token_lists = []
    counter = 0

    for article in articles:
        if counter < 10:
            title = article.get("title", "")
            link = article.get("link", "")
            content = ""
    
            if link:
                driver = webdriver.Chrome()
                driver.get(link)
                time.sleep(5)
    
                soup = BeautifulSoup(driver.page_source, "html.parser")
                paragraphs = soup.find_all("p")
                content = " ".join(p.get_text(strip=True) for p in paragraphs[16:22])
                print(content)
                driver.quit()
                
            tokens = preprocess_tokens(content)
            token_lists.append(tokens)
            processed_texts.append(" ".join(tokens))
    
            raw_data.append({
                "Cikk címe": title,
                "Cikk linkje": link,
                "Cikk tartalma": content
            })
            #counter += 1

    # Top 10 gyakori szó (stemming + lemmatizálás után, ha van ilyen lépés korábban)
    for i, article in enumerate(raw_data):
        counter = Counter(token_lists[i])
        top10_common = [word for word, _ in counter.most_common(10)]
        article["Top 10 szó"] = ", ".join(top10_common)

    # TF-IDF számítás
    vectorizer = TfidfVectorizer()
    tfidf_matrix = vectorizer.fit_transform(processed_texts)
    feature_names = vectorizer.get_feature_names_out()

    for i, article in enumerate(raw_data):
        tfidf_vector = tfidf_matrix[i]
        tfidf_scores = zip(tfidf_vector.indices, tfidf_vector.data)
        sorted_words = sorted(tfidf_scores, key=lambda x: x[1], reverse=True)
        top_words = [feature_names[idx] for idx, _ in sorted_words[:10]]
        article["Top 10 TFIDF szó"] = ", ".join(top_words)

    return raw_data


def clean_articles_sstonework(articles):
    raw_data = []
    processed_texts = []
    token_lists = []
    counter = 0

    
    for article in articles:
        if counter < 10:
            title = article.get("title", "")
            link = article.get("link", "")
            content = ""
    
            if link:
                driver = webdriver.Chrome()
                driver.get(link)
                time.sleep(5)
    
                soup = BeautifulSoup(driver.page_source, "html.parser")
                paragraphs = soup.find_all("p")
                content = " ".join(p.get_text(strip=True) for p in paragraphs[:6])
                print(content)
                driver.quit()
                
            tokens = preprocess_tokens(content)
            token_lists.append(tokens)
            processed_texts.append(" ".join(tokens))
    
            raw_data.append({
                "Cikk címe": title,
                "Cikk linkje": link,
                "Cikk tartalma": content
            })
            #counter = counter + 1

    # Top 10 common words (after stemming + lemmatization)
    for i, article in enumerate(raw_data):
        counter = Counter(token_lists[i])
        top10_common = [word for word, _ in counter.most_common(10)]
        article["Top 10 szó"] = ", ".join(top10_common)

    # TF-IDF számítás
    vectorizer = TfidfVectorizer()
    tfidf_matrix = vectorizer.fit_transform(processed_texts)
    feature_names = vectorizer.get_feature_names_out()

    for i, article in enumerate(raw_data):
        tfidf_vector = tfidf_matrix[i]
        tfidf_scores = zip(tfidf_vector.indices, tfidf_vector.data)
        sorted_words = sorted(tfidf_scores, key=lambda x: x[1], reverse=True)
        top_words = [feature_names[idx] for idx, _ in sorted_words[:10]]
        article["Top 10 TFIDF szó"] = ", ".join(top_words)

    return raw_data


def clean_articles_news(articles):
    raw_data = []
    processed_texts = []
    token_lists = []
    counter = 0

    
    for article in articles:
        if counter < 10:
            title = article.get("title", "")
            link = article.get("link", "")
            content = ""
    
            if link:
                driver = webdriver.Chrome()
                driver.get(link)
                time.sleep(5)
    
                soup = BeautifulSoup(driver.page_source, "html.parser")
                paragraphs = soup.find_all("p")
                content = " ".join(p.get_text(strip=True) for p in paragraphs[1:4])
                print(content)
                driver.quit()
                
            tokens = preprocess_tokens(content)
            token_lists.append(tokens)
            processed_texts.append(" ".join(tokens))
    
            raw_data.append({
                "Cikk címe": title,
                "Cikk linkje": link,
                "Cikk tartalma": content
            })
            #counter = counter + 1

    # Top 10 common words (after stemming + lemmatization)
    for i, article in enumerate(raw_data):
        counter = Counter(token_lists[i])
        top10_common = [word for word, _ in counter.most_common(10)]
        article["Top 10 szó"] = ", ".join(top10_common)

    # TF-IDF számítás
    vectorizer = TfidfVectorizer()
    tfidf_matrix = vectorizer.fit_transform(processed_texts)
    feature_names = vectorizer.get_feature_names_out()

    for i, article in enumerate(raw_data):
        tfidf_vector = tfidf_matrix[i]
        tfidf_scores = zip(tfidf_vector.indices, tfidf_vector.data)
        sorted_words = sorted(tfidf_scores, key=lambda x: x[1], reverse=True)
        top_words = [feature_names[idx] for idx, _ in sorted_words[:10]]
        article["Top 10 TFIDF szó"] = ", ".join(top_words)

    return raw_data


def save_to_excel():
    # Weboldalak adatainak lekérése
    market_news = news()
    celonis_news = celonis()
    sstonework_news = sstonework()
    servicenow_news = servicenow() 

    # Tisztítás és tartalom hozzáadása
    dfs = {
        "MarketResearchFuture_News": pd.DataFrame(clean_articles_news(market_news)),
        "Celonis": pd.DataFrame(clean_articles_celonis(celonis_news)),
        "SstoneWork": pd.DataFrame(clean_articles_sstonework(sstonework_news)),
        "ServiceNow": pd.DataFrame(clean_articles_servicenow(servicenow_news))
    }

    # Excel fájl létrehozása és írás
    output_file = "sites.xlsx"
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for sheet_name, df in dfs.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        writer.book.save(output_file)

    # Excel fájl betöltése diagramhoz
    wb = load_workbook(output_file)

    for sheet_name in dfs.keys():
        sheet = wb[sheet_name]
        df = dfs[sheet_name]

        def create_word_stats_and_chart(column_name, col_offset, chart_position):
            # Szavak kinyerése és gyakoriság számítás
            all_words = []
            for row in df[column_name]:
                words = [word.strip() for word in row.split(",") if word.strip()]
                all_words.extend(words)

            word_counts = Counter(all_words)
            top20 = word_counts.most_common(20)

            # Fejlécek
            header_col_word = col_offset
            header_col_freq = col_offset + 1
            sheet.cell(row=1, column=header_col_word, value=f"Top szó - {column_name}")
            sheet.cell(row=1, column=header_col_freq, value="Gyakoriság")

            for i, (word, count) in enumerate(top20, start=2):
                sheet.cell(row=i, column=header_col_word, value=word)
                sheet.cell(row=i, column=header_col_freq, value=count)

            # Diagram létrehozása
            chart = BarChart()
            chart.title = f"Top 20 szó - {column_name} alapján"
            chart.x_axis.title = "Szó"
            chart.y_axis.title = "Előfordulás"

            data = Reference(sheet, min_col=header_col_freq, min_row=1, max_row=21)
            categories = Reference(sheet, min_col=header_col_word, min_row=2, max_row=21)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.height = 10
            chart.width = 20

            sheet.add_chart(chart, chart_position)

        # Diagram a 4. oszlop (Top 10 szó) alapján — diagram az E1 helyen
        create_word_stats_and_chart("Top 10 szó", col_offset=6, chart_position="K1")

        # Diagram az 5. oszlop (Top 10 TFIDF szó) alapján — diagram az O1 helyen
        create_word_stats_and_chart("Top 10 TFIDF szó", col_offset=9, chart_position="X1")

    # Mentés
    wb.save(output_file)
    
    
def arxiv_main():
    # ArXiv AI kategória lekérdezés (max 1000 legfrissebb)
    BASE_URL = "http://export.arxiv.org/api/query?search_query=cat:cs.AI&start=0&max_results=1000"
    feed = feedparser.parse(BASE_URL)

    # Kulcsszavak, amiket vizsgálunk
    keywords = [
        # --- RETAIL ---
        "Demand forecasting",
        "Dynamic pricing",
        "Recommendation system",
        "Customer segmentation",
        "Chatbot",
        "Conversational commerce",
        "Generative AI for marketing",
        "Personalized marketing",
        "Loyalty prediction",
        "Inventory optimization",
        "Omnichannel",
        "Store analytics",
        "In-store tracking",
        "Foot traffic prediction",
        "Heatmap analytics",
        "Personalization",
        "Shopping assistant",
        "A/B testing",
        "LLM for e-commerce",
        "Multimodal product search",
        "Visual search",
        "Voice assistant",
        "Product tagging",
        "Fraud detection",
        "Supply chain AI",
        "Delivery route optimization",
        "Returns prediction",
        "Shelf monitoring",
        "Planogram compliance",
        "Dynamic ad generation",
        "Price elasticity modeling",
        "Churn prediction",
        "Sentiment analysis",
        "Social media monitoring",
        "Computer vision for retail",
        "Smart shelves",
        "Automated checkout",
        "RFID-based analytics",
        "Digital twin retail",
        "Stockout prediction",

        # --- SSC (Shared Services Center) ---
        "Document automation",
        "Invoice processing",
        "AI in finance",
        "Process mining",
        "Process optimization",
        "RPA",
        "Intelligent document processing",
        "OCR",
        "IDP",
        "NLP for back office",
        "AI chatbot for HR",
        "Workflow automation",
        "Predictive analytics",
        "Self-service portal",
        "Semantic search",
        "Contract analysis",
        "Knowledge management",
        "Time tracking automation",
        "Email triage",
        "Email intent detection",
        "SLA prediction",
        "LLM for operations",
        "Task prioritization",
        "Internal ticket classification",
        "Anomaly detection in finance",
        "Vendor matching",
        "Procure-to-pay automation",
        "Order-to-cash optimization",
        "Expense report classification",
        "Employee onboarding automation",
        "Voice-to-text for meetings",
        "HR query resolution",
        "Multi-language support",
        "Budget forecasting",
        "Cash flow prediction",
        "Legal document summarization",
        "Data entry automation",
        "Audit risk detection",
        "Policy compliance check",
        "Meeting summarization",
        "Data harmonization",
        "Low-code AI integration",
        "Decision intelligence",
        "Reconciliation automation"
    ]


    # Számoljuk, melyik szó hányszor szerepel a címekben és kivonatokban
    counter = Counter()
    for entry in tqdm(feed.entries, desc="Publikációk feldolgozása"):
        content = (entry.title + " " + entry.summary).lower()
        for kw in keywords:
            if kw.lower() in content:
                counter[kw] += 1

    # Eredmény előkészítése
    labels, values = zip(*counter.most_common())

    # Vizualizáció
    sns.set(style="whitegrid")
    plt.figure(figsize=(12, 7))
    sns.barplot(x=list(values), y=list(labels), palette="coolwarm")
    plt.title("AI Trendek az arXiv-on (1000 legfrissebb AI publikáció)", fontsize=16)
    plt.xlabel("Előfordulások száma")
    plt.ylabel("Kulcsszavak")
    plt.tight_layout()

    plt.savefig("ai_trends_arxiv.png")

    # Készítsük el a DataFrame-t és mentsük el Excel fájlba
    df = pd.DataFrame(list(zip(labels, values)), columns=["Keyword", "Occurrences"])
    df.to_excel("ai_trends_arxiv.xlsx", index=False)

    # Openpyxl betöltése az Excel fájlhoz
    wb = load_workbook("ai_trends_arxiv.xlsx")
    ws = wb.active

    ws.title = "arXiv"

    # Kép hozzáadása a D oszlop mellé
    img = Image("ai_trends_arxiv.png")
    img.anchor = 'E1'  # A kép az E1 cellába kerül, tehát a D oszloptól jobbra

    # Kép hozzáadása az Excel fájlhoz
    ws.add_image(img)

    # Mentés az új fájlba
    wb.save("arxiv.xlsx")
    print("Sikeresen elkészültek az output fájlok!")

# Használat:
save_to_excel()
arxiv_main()



