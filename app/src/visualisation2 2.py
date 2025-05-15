import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import numpy as np
import seaborn as sns

# Beolvasás
data = pd.read_excel("demo2.xlsx")

# ExcelWriter példány létrehozása
writer = pd.ExcelWriter("output3.xlsx", engine='openpyxl')

# Első munkalap: adatok
data.to_excel(writer, sheet_name='data', index=False)

# Második munkalap: üres vizualizációs lap
pd.DataFrame().to_excel(writer, sheet_name='visualisation', index=False)

# Töltsük be a financial_data_with_ai_buzz_score.xlsx fájlt
financial_data = pd.read_excel("financial_data_with_buzz.xlsx")

# Hozzunk létre egy új munkalapot 'financial_data' néven, és töltsük be a pénzügyi adatokat
financial_data.to_excel(writer, sheet_name='financial_data', index=False)

# Kördiagram létrehozása a "number of news" oszlop alapján
news_counts = data['Number of news'].value_counts().sort_index()

# Még lágyabb színek palettája (pasztell színek)
colors = [
    '#F9E0E0',  # Nagyon halvány pasztell rózsaszín
    '#E7F9E7',  # Nagyon halvány pasztell zöld
    '#FFFCE1',  # Nagyon halvány pasztell sárga
    '#E9D9FF',  # Nagyon halvány pasztell lila
    '#E7F6FF',  # Nagyon halvány pasztell kék
    '#FFF1F4',  # Nagyon halvány pasztell barack
    '#D9F2FF',  # Nagyon halvány pasztell égkék
    '#FBEAEA',  # Nagyon halvány pasztell piros
    '#F8E8D3',  # Nagyon halvány pasztell homokszín
    '#F6E8FF'   # Nagyon halvány pasztell levendula
]

# Kördiagram rajzolása
fig, ax = plt.subplots(figsize=(6, 6))  # Kördiagram különböző helyen
ax.pie(news_counts, labels=news_counts.index, autopct='%1.1f%%', startangle=90, colors=colors)
ax.axis('equal')  # A kör alak megőrzése
ax.set_title("Distribution of 'Number of News'", fontsize=16)

# Kördiagram mentése külön fájlba
plt.savefig("pie_chart.png", format="png")
plt.close()  # Bezárjuk a plt-t a következő diagram előtt

# Hisztogram rajzolása
fig, ax = plt.subplots(figsize=(6, 6))  # Hisztogram különböző helyen
ax.hist(data['AI buzz score'], bins=10, color='#F0D8D8', edgecolor='black')

# Hisztogram X tengelyének beállítása minimumtól maximumig
min_value = data['AI buzz score'].min()
max_value = data['AI buzz score'].max()
ax.set_xlim(min_value, max_value)  # X tengely beállítása az adat minimum és maximum értékei között

ax.set_title("Distribution of AI Buzz Score", fontsize=16)
ax.set_xlabel("AI Buzz Score")
ax.set_ylabel("Frequency")

# Hisztogram mentése külön fájlba
plt.savefig("histogram.png", format="png")
plt.close()  # Bezárjuk a plt-t

# Képek beillesztése az Excel fájlba
img_pie = Image("pie_chart.png")
img_histogram = Image("histogram.png")

# Excel munkalap hozzáadása
sheet = writer.book["visualisation"]

# Az Excel-ben a képek elhelyezése
sheet.add_image(img_pie, 'B2')  # Kördiagram a B2 cellában (B2:G24 terjedelmen)
sheet.add_image(img_histogram, 'I2')  # Hisztogram az I2 cellában (I2:O24 terjedelmen)

# Oszlopdiagram az 'Average social media sentiment' oszlopból
fig, ax = plt.subplots(figsize=(12, 8))  # Magasabb diagram a tengelyfeliratoknak

# Csak az első szó megjelenítése a "Company" nevekből
companies = [company.split()[0] for company in data['Company']]  # Az első szó kiválasztása

# Oszlopok színének beállítása: halvány rózsaszín, ha negatív, halvány menta, ha pozitív
colors = ['#D1ECFF' if val < 0 else '#F0D8D8' for val in data['Average social media sentiment']]  # Pasztell rózsaszín és menta zöld

# Oszlopdiagram rajzolása
ax.bar(companies, data['Average social media sentiment'], color=colors)

# Tengely beállításai
ax.set_xlabel("Company")
ax.set_ylabel("Average Social Media Sentiment")
ax.set_title("Average Social Media Sentiment by Company", fontsize=16)
ax.axhline(0, color='black', linewidth=1)  # 0-ás vonal a tengelyen (negatív értékek is megjelenhetnek)

# Tengelyfeliratok függőlegesen
plt.xticks(rotation=90)  # A tengelyfeliratok elforgatása függőlegesen

# Oszlopdiagram mentése
plt.savefig("bar_chart.png", format="png")
plt.close()  # Bezárjuk a plt-t

# Oszlopdiagram hozzáadása az Excel fájlhoz
img_bar = Image("bar_chart.png")

# Az Excel-ben az oszlopdiagram elhelyezése
sheet.add_image(img_bar, 'B26')  # Az oszlopdiagram a B26 cellába kerül (B26:T42)


# **Új oszlopdiagram az 'Average news sentiment' oszlopból**
fig, ax = plt.subplots(figsize=(12, 8))  # Magasabb diagram a tengelyfeliratoknak

# Csak az első szó megjelenítése a "Company" nevekből
companies = [company.split()[0] for company in data['Company']]  # Az első szó kiválasztása

# Oszlopok színének beállítása: halvány rózsaszín, ha negatív, halvány menta, ha pozitív
colors = ['#D1ECFF' if val < 0 else '#F0D8D8' for val in data['Average news sentiment']]  # Pasztell rózsaszín és menta zöld

# Oszlopdiagram rajzolása
ax.bar(companies, data['Average news sentiment'], color=colors)

# Tengely beállításai
ax.set_xlabel("Company")
ax.set_ylabel("Average News Sentiment")
ax.set_title("Average News Sentiment by Company", fontsize=16)
ax.axhline(0, color='black', linewidth=1)  # 0-ás vonal a tengelyen (negatív értékek is megjelenhetnek)

# Tengelyfeliratok függőlegesen
plt.xticks(rotation=90)  # A tengelyfeliratok elforgatása függőlegesen

# Oszlopdiagram mentése
plt.savefig("bar_chart_news_sentiment.png", format="png")
plt.close()  # Bezárjuk a plt-t

# Oszlopdiagram hozzáadása az Excel fájlhoz
img_bar_news_sentiment = Image("bar_chart_news_sentiment.png")

# Az Excel-ben az oszlopdiagram elhelyezése
sheet.add_image(img_bar_news_sentiment, 'O26')  # Az oszlopdiagram a P26 cellába kerül (P26:O42)


def create_merged_cell(sheet, merge_range, cell_address, text, color):
    """
    Létrehoz egy összevont cellát, beállítja a szöveget, középre igazítja, félkövér betűtípust alkalmaz,
    fekete szegélyt ad hozzá, és egyéni háttérszínt alkalmaz.

    :param sheet: Az Excel munkalap, ahol a módosításokat végezzük
    :param merge_range: Az összevont cellák tartománya (pl. 'U5:X5')
    :param cell_address: A cella címének megadása, ahol a szöveg lesz (pl. 'P5')
    :param text: A beírt szöveg (pl. 'Number of News')
    :param color: A háttérszín hex kódja (pl. "E0F7FA" a világoskékhez)
    """
    # Összevont cellák létrehozása
    sheet.merge_cells(merge_range)

    # Cella kiválasztása
    cell = sheet[cell_address]
    cell.value = text  # Szöveg beállítása

    # A szöveg középre igazítása és félkövér betűtípus beállítása
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)

    # Egyéni háttérszín beállítása
    custom_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.fill = custom_fill

# Példa a használatra:
create_merged_cell(sheet, 'Q2:S3', 'Q2', 'Average', "D1ECFF")  # Világoskék háttér
create_merged_cell(sheet, 'U2:W3', 'U2', 'ALDI', "F0D8D8")   
create_merged_cell(sheet, 'U5:W5', 'U5', 'Number of News', "F0D8D8")  # Halvány rózsaszín
create_merged_cell(sheet, 'Q5:S5', 'Q5', 'Number of News', "D1ECFF")  
create_merged_cell(sheet, 'Q6:S7', 'Q6', str(round(np.mean(data['Number of news']), 2)), "F9F9F9")  
create_merged_cell(sheet, 'U6:W7', 'U6', str(round(data.loc[data['Company'] == 'ALDI', 'Number of news'].values[0], 2)), "F9F9F9")
create_merged_cell(sheet, 'U9:W9', 'U9', 'Average news sentiment', "F0D8D8")  # Halvány rózsaszín
create_merged_cell(sheet, 'Q9:S9', 'Q9', 'Average news sentiment', "D1ECFF")  
create_merged_cell(sheet, 'Q10:S11', 'Q10', str(round(np.mean(data['Average news sentiment']), 2)), "F9F9F9")  
create_merged_cell(sheet, 'U10:W11', 'U10', str(round(data.loc[data['Company'] == 'ALDI', 'Average news sentiment'].values[0], 2)), "F9F9F9")
create_merged_cell(sheet, 'U13:W13', 'U13', 'Number of AI key words', "F0D8D8")  # Halvány rózsaszín
create_merged_cell(sheet, 'Q13:S13', 'Q13', 'Number of AI key words', "D1ECFF")  
create_merged_cell(sheet, 'Q14:S15', 'Q14', str(round(np.mean(data['Number of AI key words']), 2)), "F9F9F9")  
create_merged_cell(sheet, 'U14:W15', 'U14', str(round(data.loc[data['Company'] == 'ALDI', 'Number of AI key words'].values[0], 2)), "F9F9F9")
create_merged_cell(sheet, 'U17:W17', 'U17', 'Average social media sentiment', "F0D8D8")  # Halvány rózsaszín
create_merged_cell(sheet, 'Q17:S17', 'Q17', 'Average social media sentiment', "D1ECFF")  
create_merged_cell(sheet, 'Q18:S19', 'Q18', str(round(np.mean(data['Average social media sentiment']), 2)), "F9F9F9")  
create_merged_cell(sheet, 'U18:W19', 'U18', str(round(data.loc[data['Company'] == 'ALDI', 'Average social media sentiment'].values[0], 2)), "F9F9F9")
create_merged_cell(sheet, 'U21:W21', 'U21', 'AI buzz score', "F0D8D8")  # Halvány rózsaszín
create_merged_cell(sheet, 'Q21:S21', 'Q21', 'AI buzz score', "D1ECFF")  
create_merged_cell(sheet, 'Q22:S23', 'Q22', str(round(np.mean(data['AI buzz score']), 2)), "F9F9F9")  
create_merged_cell(sheet, 'U22:W23', 'U22', str(round(data.loc[data['Company'] == 'ALDI', 'AI buzz score'].values[0], 2)), "F9F9F9")

# Pasztell színpaletta a diagramhoz
colors_stronger = [
    '#B3DDF2',  # Világos pasztellkék
    '#7FC9F1',  # Középkék, világos, de élénk
    '#4FB0E6',  # Erőteljesebb kék
    '#2C98D1'   # Mélyebb, de még mindig világosabb kék
]

# Csoportosított adatok a 'Profit Rate (%)' alapján
df_grouped = financial_data.groupby(["Company name", "Year"]).agg({"Profit Rate (%)": "mean"}).reset_index()

# Ábra előkészítése
plt.figure(figsize=(14, 8))

# Barplot elkészítése a profitráta éves változásával
sns.barplot(
    x="Company name",
    y="Profit Rate (%)",
    hue="Year",
    data=df_grouped,
    palette=colors_stronger
)

# Set chart title and axis labels
plt.title("Profit Rate Annual Change by Company", fontsize=16)
plt.xlabel("Company", fontsize=14)
plt.ylabel("Profit Rate (%)", fontsize=14)

# Rotate X-axis labels for readability
plt.xticks(rotation=90)

# Add legend
plt.legend(title='Year', loc='upper left')

# Adjust layout and save the plot
plt.tight_layout()
plt.savefig("profit_rate_by_company_year.png", format="png")
plt.close()

# Insert image into the Excel sheet
img_profit_rate = Image("profit_rate_by_company_year.png")
sheet = writer.sheets["visualisation"]
sheet.add_image(img_profit_rate, 'B60')

# Create scatter plot: AI buzz score vs Profit Rate
df_grouped2 = financial_data.groupby(["Company name"]).agg({
    "Profit Rate (%)": "mean",
    "AI buzz score": "mean"
}).reset_index()

plt.figure(figsize=(10, 6))

x = df_grouped2['AI buzz score']
y = df_grouped2['Profit Rate (%)']

plt.scatter(x, y, alpha=0.5, c=x, cmap='viridis', edgecolors='w', linewidth=0.5)

plt.title("AI Buzz Score vs Profit Rate", fontsize=16)
plt.xlabel("AI Buzz Score", fontsize=14)
plt.ylabel("Profit Rate (%)", fontsize=14)

plt.tight_layout()
plt.savefig("ai_buzz_vs_profit_rate.png", format="png")
plt.close()

# Insert updated image into Excel sheet
sheet = writer.sheets["visualisation"]
img_ai_buzz_vs_profit_rate = Image("ai_buzz_vs_profit_rate.png")
sheet.add_image(img_ai_buzz_vs_profit_rate, 'R60')

# Update merged cell with title and correlation value
create_merged_cell(sheet, 'T84:V88', 'T84', 'Correlation between AI buzz score and Profit Rate', "D1ECFF")
cell = sheet['T84']
cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell.font = Font(bold=True)
cell.fill = PatternFill(start_color="D1ECFF", end_color="D1ECFF", fill_type="solid")

correlation = df_grouped2['Profit Rate (%)'].corr(df_grouped2['AI buzz score'])
create_merged_cell(sheet, 'W84:Y88', 'W84', str(round(correlation, 2)), "F9F9F9")

# Save workbook
writer.close()
